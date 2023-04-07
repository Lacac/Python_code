# Convert .json file ---> .csv file & .xlsx file

import json
import csv

# json ---> csv


input_json_file='jsonTest.json'
output_csv_file='csv_file.csv'
input = open(input_json_file)
data = json.load(input)
print(type(data))

input.close()
output = csv.writer(open('output_csv_file.csv','w'))
output.writerow(data[0].keys())                 # header row
for row in data:
    output.writerow(row.values())




# jason --> excel

import openpyxl

input_json_file='jsonTest.json'
input = open(input_json_file)
data = json.load(input)

wb = openpyxl.Workbook()
sheet = wb.active
numberOfTitle = len(data[0].keys())
listOfTitle = list()
for i in data[0]:
    listOfTitle.append(i)

# print(listOfTitle)
for i in range(len(listOfTitle)):
    sheet.cell(row = 1, column = i+1).value = listOfTitle[i]

# print(list(data[0].values())[0])


currentRow = 2
currentIndexOfData = 0     
for row in data:
    currentData = list(data[currentIndexOfData].values())
    sheet.cell(row = currentRow, column = 1).value = currentData[0]
    sheet.cell(row = currentRow, column = 2).value = currentData[1]
    currentRow += 1

wb.save('Ass_Day5_U2.xlsx')



