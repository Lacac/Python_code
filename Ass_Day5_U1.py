from collections import namedtuple
from tabnanny import check
import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import csv
import datetime

# Request the first link: 
url = 'https://www.us-cert.gov/ncas/alerts'
res = requests.get(url)
try:
    res.raise_for_status()
except: 
    print('ERROR')
mySoup = BeautifulSoup(res.text, 'html.parser')

# Khởi tạo bảng kết quả: 
wb = openpyxl.Workbook()
# wb.save('Alert.xlsx')
sheet = wb.active
sheet['A1'] = 'ID'
sheet['B1'] = 'Name'
sheet['C1'] = 'Release date'
sheet['D1'] = 'Revise date'
sheet['E1'] = 'Link'
sheet['F1'] = 'Tips'

sheet.column_dimensions['A'] = 15
sheet.column_dimensions['B'] = 70
sheet.column_dimensions['C'] = 20
sheet.column_dimensions['D'] = 20
sheet.column_dimensions['E'] = 70

# Function : fill the information to excel:
def fillInforToExcel(sheet,currentRow, id, name, releaseDate, lastRevisedDate, link): 
    sheet.cell(row = currentRow, column = 1).value = id
    sheet.cell(row = currentRow, column = 2).value = name
    sheet.cell(row = currentRow, column = 3).value = releaseDate
    sheet.cell(row = currentRow, column = 4).value = lastRevisedDate
    sheet.cell(row = currentRow, column = 5).value = link
    sheet.cell(row = currentRow, column = 6).value = ''

# create csv file: 
outputFile = open('Ass_Day5_U1.csv','w',newline ='')
outputDicWriter = csv.DictWriter(outputFile,['ID','Name','Release date','Revise date','Link','Tips'],delimiter='\t', lineterminator='\n\n')

def fillInforToCsv(outputDicWriter,id,name,releaseDate,reviseDate,link):
    outputDicWriter.writerow({'ID':id,'Name':name,'Release date':releaseDate,'Revise date':reviseDate,'Link':link,'Tips':''})


# Get all the page: 
pageTag = mySoup.select('.pager__item')                 # class = paper__item
setPage = set()
for page in pageTag:
    nextPage = page.select('a')[0].get('href')          # get the link of page 
    setPage.add(nextPage)                               # add the link of page to set()
print(setPage)

listPage = list()
for page in setPage:
    listPage.append(page)
listPage.sort()
print(listPage)


# Xử lý chính:
currentYear = str(datetime.datetime.today().year)
idYear = currentYear[len(currentYear)-2:]


# Function: check the ID match the current year
def checkID(id, idYear):
    regex = re.compile(r'[A-Z]+'+ re.escape(idYear)+r'-')
    check = regex.match(id)
    if check != None:
        return True
    return False

currentRow = 2
for page in listPage:
    linkNextPage = 'https://www.cisa.gov/uscert/ncas/alerts' + page
    res = requests.get(linkNextPage)
    try: 
        res.raise_for_status()
    except:
        print('ERROR')

    mySoup = BeautifulSoup(res.text,'html.parser')

    issuedTags = mySoup.select('.views-field.views-field-title')  # link issue: 'https://www.cisa.gov' + /ncas/alerts/aa22-223a
                                                                    # https://www.cisa.gov/uscert/ncas/alerts/aa22-228a
    for issuedTag in issuedTags:      
        link = 'https://www.cisa.gov/uscert' + issuedTag.select('a')[0].get('href')
        id = str(issuedTag.select('span')[0].contents[0]).strip()
        id = id[:(len(id)-1)]
        if check(id,idYear) == False:
            continue
        name = str(issuedTag.select('a')[0].contents[0])
        res = requests.get(link)
        issuedSoup = BeautifulSoup(res.text,'html.parser')
        print('LINK: ',link)
        date = issuedSoup.select('.submitted.meta-text')
        print('DATE: ', date)
    #    print(page,issuedTag, date)
    #    print(date)
    #    print(issuedTag)
        dateString = date[0].contents
        releaseDate = str(dateString[0])            # release date, chưa biết có dấu | hay không

        # TH1: 2 date khác nhau, tức: releaseDate: có dang ....| Last, len(dateString) = 3
        if len(dateString) == 3:                    # Original release date: July 06, 2022  | Last ', <a href="#revisions">revised</a>, ': July 0
            regex1 = re.compile(r'.*Original release date:([^|]*)| Last')
            releaseDate = regex1.search(releaseDate).group(1).strip()
        
            lastRevisedDate = str(dateString[2])
            regex2 = re.compile(r': (.*)')
            lastRevisedDate = regex2.match(lastRevisedDate).group(1).strip()
        # TH2 : 2 date = nhau
        elif len(dateString) == 1:                   # release = last revised
            regexNoRevise = re.compile(r'.*Original release date:(.*)') 
            releaseDate = regexNoRevise.search(releaseDate).group(1).strip()
            lastRevisedDate = releaseDate

        fillInforToCsv(outputDicWriter,id,name,releaseDate,lastRevisedDate,link)

        fillInforToExcel(wb.active,currentRow,id,name,releaseDate,lastRevisedDate,link)
        currentRow += 1
wb.save('Test3.xlsx')
outputFile.close()


       
